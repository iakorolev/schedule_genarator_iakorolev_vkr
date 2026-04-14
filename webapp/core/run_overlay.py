"""Хранение и применение overlay-правок поверх исходного файла РУН."""

from __future__ import annotations

import json
import shutil
from collections import Counter
from copy import copy
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter

from core.un_parser import pick_un_sheet

OVERLAY_FILENAME = "run_overlay.json"
EDITED_RUN_FILENAME = "run_edited.xlsx"
EFFECTIVE_RUN_FILENAME = "run_effective.xlsx"
RAW_SHEET_CACHE_FILENAME = "run_editor_rows.json"
DELETED_FLAG = "__deleted__"
ADDED_ROWS_KEY = "added_rows"
NEXT_ADDED_ID_KEY = "next_added_id"
ADDED_ROW_PREFIX = "new:"

HEADER_ROWS = 5
DATA_START_ROW = 6

FIELD_SPECS: list[tuple[str, str]] = [
    ("Код_ООП", "Код Направление /специальность /образовательная программа"),
    ("ООП", "Образовательная программа"),
    ("Дисциплина", "Наименование дисциплины или вида учебной работы"),
    ("Семестр", "Семестр"),
    ("Вид_работы", "Вид учебной работы"),
    ("Код_группы", "Учебная группа"),
    ("Номер_группы", "Номер группы"),
    ("Количество_студентов", "Кол-во чел. в группе (потоке) Всего"),
    ("Кафедра", "Сведения о ППС Кафедра"),
    ("Должность", "должность"),
    ("Тип_занятости", "штатн."),
    ("Преподаватель", "Фамилия И.О. преподавателя"),
    ("Лекции_часы", "Объём учебной работы ППС Лекции"),
    ("Практика_часы", "Практика / Семинары"),
    ("Лабораторные_часы", "Лаб. работы / Клинические занятия"),
    ("Всего_часов", "Всего часов"),
]

EDITABLE_FIELDS = [
    "Дисциплина",
    "Вид_работы",
    "Код_группы",
    "Номер_группы",
    "Преподаватель",
    "Кафедра",
    "Семестр",
    "Лекции_часы",
    "Практика_часы",
    "Лабораторные_часы",
    "Всего_часов",
]

VIEW_FIELDS = [
    "Преподаватель",
    "Дисциплина",
    "Вид_работы",
    "Код_группы",
    "Номер_группы",
    "Кафедра",
    "Семестр",
    "Лекции_часы",
    "Практика_часы",
    "Лабораторные_часы",
    "Всего_часов",
]

NEW_ROW_EXTRA_FIELDS = [
    "Код_ООП",
    "ООП",
    "Количество_студентов",
    "Должность",
    "Тип_занятости",
]

NEW_ROW_FIELDS = VIEW_FIELDS + NEW_ROW_EXTRA_FIELDS

SUGGEST_FIELDS = [
    "Преподаватель",
    "Дисциплина",
    "Вид_работы",
    "Код_группы",
    "Номер_группы",
    "Кафедра",
    "Семестр",
    "Должность",
    "Тип_занятости",
    "ООП",
    "Код_ООП",
]
MAX_SUGGESTIONS_PER_FIELD = 200

NUMERIC_FIELDS = {
    "Лекции_часы",
    "Практика_часы",
    "Лабораторные_часы",
    "Всего_часов",
    "Количество_студентов",
}

DELETE_CLEAR_FIELDS = [
    "Дисциплина",
    "Вид_работы",
    "Код_группы",
    "Номер_группы",
    "Преподаватель",
    "Кафедра",
    "Семестр",
    "Лекции_часы",
    "Практика_часы",
    "Лабораторные_часы",
    "Всего_часов",
]


def _normalize_header_series(excel_data: pd.DataFrame) -> pd.Series:
    """Склеивает верхние строки заголовка в единый набор имён колонок."""
    header_rows = excel_data.iloc[:HEADER_ROWS]
    combined = header_rows.fillna("").astype(str).agg(" ".join)
    return combined.str.replace(r"\s+", " ", regex=True).str.strip()


def _normalize_cell_value(value: Any) -> str:
    """Приводит значение ячейки к строке для JSON и сравнения."""
    if value is None:
        return ""
    if isinstance(value, float):
        if pd.isna(value):
            return ""
        if value.is_integer():
            return str(int(value))
        return str(value).rstrip("0").rstrip(".") if "." in str(value) else str(value)
    if isinstance(value, int):
        return str(value)
    text = str(value).strip()
    if text.lower() == "nan":
        return ""
    return text


def _is_truthy(value: Any) -> bool:
    return str(value).strip().lower() in {"1", "true", "yes", "on", "y"}


def _overlay_row_deleted(row_overlay: dict[str, Any] | None) -> bool:
    if not isinstance(row_overlay, dict):
        return False
    return _is_truthy(row_overlay.get(DELETED_FLAG, ""))


def _cell_value_for_deleted(field: str) -> Any:
    if field in NUMERIC_FIELDS:
        return 0
    return ""


def _overlay_path(session_dir: Path) -> Path:
    return session_dir / OVERLAY_FILENAME


def _edited_run_path(session_dir: Path) -> Path:
    return session_dir / EDITED_RUN_FILENAME


def _effective_materialized_run_path(session_dir: Path) -> Path:
    return session_dir / EFFECTIVE_RUN_FILENAME


def _sheet_cache_path(session_dir: Path) -> Path:
    return session_dir / RAW_SHEET_CACHE_FILENAME


def _empty_overlay() -> dict[str, Any]:
    return {"rows": {}, ADDED_ROWS_KEY: {}, NEXT_ADDED_ID_KEY: 1}


def _normalize_overlay_record(value: Any) -> dict[str, str]:
    if not isinstance(value, dict):
        return {}
    return {str(k): _normalize_cell_value(v) for k, v in value.items()}


def _is_added_row_id(row_id: Any) -> bool:
    return str(row_id).startswith(ADDED_ROW_PREFIX)


def _normalize_row_id(row_id: Any) -> str:
    if row_id is None:
        raise ValueError("row_id is required")
    text = str(row_id).strip()
    if not text:
        raise ValueError("row_id is required")
    if _is_added_row_id(text):
        return text
    try:
        return str(int(float(text)))
    except Exception as exc:
        raise ValueError(f"invalid row_id: {row_id}") from exc


def init_overlay_store(session_dir: Path) -> None:
    """Создаёт пустое хранилище overlay-правок."""
    save_overlay(session_dir, _empty_overlay())


def load_overlay(session_dir: Path) -> dict[str, Any]:
    """Читает JSON с overlay-правками."""
    path = _overlay_path(session_dir)
    if not path.exists():
        return _empty_overlay()
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return _empty_overlay()

    rows = data.get("rows") if isinstance(data, dict) else None
    added_rows = data.get(ADDED_ROWS_KEY) if isinstance(data, dict) else None
    next_added_id = data.get(NEXT_ADDED_ID_KEY) if isinstance(data, dict) else None

    clean_rows: dict[str, dict[str, str]] = {}
    if isinstance(rows, dict):
        for key, value in rows.items():
            clean_rows[str(key)] = _normalize_overlay_record(value)

    clean_added_rows: dict[str, dict[str, str]] = {}
    if isinstance(added_rows, dict):
        for key, value in added_rows.items():
            row_id = str(key).strip()
            if not row_id:
                continue
            if not _is_added_row_id(row_id):
                row_id = f"{ADDED_ROW_PREFIX}{row_id}"
            clean_added_rows[row_id] = _normalize_overlay_record(value)

    try:
        next_added_id = max(1, int(next_added_id or 1))
    except Exception:
        next_added_id = 1

    for row_id in clean_added_rows:
        try:
            suffix = int(str(row_id).split(":", 1)[1])
            next_added_id = max(next_added_id, suffix + 1)
        except Exception:
            continue

    return {
        "rows": clean_rows,
        ADDED_ROWS_KEY: clean_added_rows,
        NEXT_ADDED_ID_KEY: next_added_id,
    }


def save_overlay(session_dir: Path, overlay: dict[str, Any]) -> None:
    """Сохраняет JSON с overlay-правками."""
    rows = overlay.get("rows") if isinstance(overlay, dict) else {}
    added_rows = overlay.get(ADDED_ROWS_KEY) if isinstance(overlay, dict) else {}
    next_added_id = overlay.get(NEXT_ADDED_ID_KEY, 1) if isinstance(overlay, dict) else 1

    payload = {
        "rows": rows if isinstance(rows, dict) else {},
        ADDED_ROWS_KEY: added_rows if isinstance(added_rows, dict) else {},
        NEXT_ADDED_ID_KEY: int(next_added_id or 1),
    }
    path = _overlay_path(session_dir)
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(path)


def clear_overlay_artifacts(session_dir: Path) -> None:
    """Удаляет производные файлы overlay, чтобы не было несогласованности."""
    for path in [_edited_run_path(session_dir), _effective_materialized_run_path(session_dir), _sheet_cache_path(session_dir)]:
        if path.exists():
            path.unlink()


def get_overlay_stats(session_dir: Path) -> dict[str, Any]:
    """Возвращает краткую статистику по правкам."""
    overlay = load_overlay(session_dir)
    rows = overlay.get("rows", {})
    added_rows = overlay.get(ADDED_ROWS_KEY, {})
    field_count = 0
    deleted_rows = 0

    for value in list(rows.values()) + list(added_rows.values()):
        if not isinstance(value, dict):
            continue
        if _overlay_row_deleted(value):
            deleted_rows += 1
        field_count += sum(1 for k in value.keys() if k != DELETED_FLAG)

    return {
        "changed_rows": len(rows) + len(added_rows),
        "changed_fields": field_count,
        "deleted_rows": deleted_rows,
        "added_rows": len(added_rows),
        "has_overlay": bool(rows or added_rows),
        "edited_run_exists": _edited_run_path(session_dir).exists(),
    }


def _find_columns(run_xlsx: Path) -> tuple[str, dict[str, dict[str, Any]], pd.DataFrame]:
    """Находит лист и реальные Excel-колонки по ожидаемым полям."""
    sheet = pick_un_sheet(run_xlsx)
    excel_data = pd.read_excel(run_xlsx, sheet_name=sheet, header=None)
    headers = _normalize_header_series(excel_data)

    columns: dict[str, dict[str, Any]] = {}
    for field_name, pattern in FIELD_SPECS:
        found_idx = None
        for idx, header in enumerate(headers.tolist()):
            if pattern in str(header):
                found_idx = idx
                break
        if found_idx is None:
            continue
        columns[field_name] = {
            "field": field_name,
            "header": str(headers.iloc[found_idx]),
            "df_index": found_idx,
            "excel_col": found_idx + 1,
        }
    return sheet, columns, excel_data


def _build_search_text(current_values: dict[str, str], deleted: bool, added: bool) -> str:
    searchable = " | ".join(current_values.get(field, "") for field in VIEW_FIELDS).lower()
    if deleted:
        searchable += " | удалено | deleted | disabled"
    if added:
        searchable += " | добавлено | added | new"
    return searchable


def _default_new_row_values(run_xlsx: Path) -> dict[str, str]:
    rows = read_run_rows(run_xlsx)
    defaults = {field: "" for field in NEW_ROW_FIELDS}
    for field in ["Кафедра", "Должность", "Тип_занятости", "ООП", "Код_ООП", "Семестр"]:
        for row in rows:
            value = _normalize_cell_value(row.get("current", {}).get(field, ""))
            if value:
                defaults[field] = value
                break
    return defaults


def get_run_field_suggestions(run_xlsx: Path, session_dir: Path | None = None) -> dict[str, list[str]]:
    """Возвращает варианты автоподсказок по значениям из РУН и overlay."""
    rows = read_run_rows(run_xlsx, session_dir)
    counters: dict[str, Counter[str]] = {field: Counter() for field in SUGGEST_FIELDS}

    for row in rows:
        if row.get("deleted"):
            continue
        current = row.get("current", {}) if isinstance(row, dict) else {}
        if not isinstance(current, dict):
            continue
        for field in SUGGEST_FIELDS:
            value = _normalize_cell_value(current.get(field, ""))
            if value:
                counters[field][value] += 1

    out: dict[str, list[str]] = {}
    for field in SUGGEST_FIELDS:
        ranked = sorted(counters[field].items(), key=lambda item: (-item[1], item[0].lower()))
        out[field] = [value for value, _ in ranked[:MAX_SUGGESTIONS_PER_FIELD]]
    return out


def read_run_rows(run_xlsx: Path, session_dir: Path | None = None) -> list[dict[str, Any]]:
    """Возвращает строки РУН в виде JSON-совместимого списка."""
    sheet, columns, excel_data = _find_columns(run_xlsx)
    overlay = load_overlay(session_dir) if session_dir else _empty_overlay()
    overlay_rows = overlay.get("rows", {})
    overlay_added_rows = overlay.get(ADDED_ROWS_KEY, {})

    data = excel_data.iloc[HEADER_ROWS:].copy().reset_index(drop=True)
    rows: list[dict[str, Any]] = []
    for idx in range(len(data)):
        excel_row = DATA_START_ROW + idx
        row_key = str(excel_row)
        original_values: dict[str, str] = {}
        current_values: dict[str, str] = {}
        changed_fields: list[str] = []
        row_overlay = overlay_rows.get(row_key, {})
        deleted = _overlay_row_deleted(row_overlay)

        for field in NEW_ROW_FIELDS:
            spec = columns.get(field)
            raw_value = ""
            if spec is not None:
                raw_value = _normalize_cell_value(data.iat[idx, spec["df_index"]])
            original_values[field] = raw_value
            current_value = _normalize_cell_value(row_overlay.get(field, raw_value))
            current_values[field] = current_value
            if field in VIEW_FIELDS and current_value != raw_value:
                changed_fields.append(field)

        row = {
            "row_id": row_key,
            "excel_row": excel_row,
            "display_row": str(excel_row),
            "sheet": sheet,
            "current": current_values,
            "original": original_values,
            "changed_fields": changed_fields,
            "changed": bool(changed_fields) or deleted,
            "deleted": deleted,
            "added": False,
            "search_text": _build_search_text(current_values, deleted=deleted, added=False),
        }
        rows.append(row)

    for row_id, row_overlay in sorted(overlay_added_rows.items(), key=lambda x: x[0]):
        if not isinstance(row_overlay, dict):
            continue
        current_values = {field: _normalize_cell_value(row_overlay.get(field, "")) for field in NEW_ROW_FIELDS}
        changed_fields = [field for field in VIEW_FIELDS if current_values.get(field, "")]
        deleted = _overlay_row_deleted(row_overlay)
        row = {
            "row_id": row_id,
            "excel_row": None,
            "display_row": row_id,
            "sheet": sheet,
            "current": current_values,
            "original": {field: "" for field in NEW_ROW_FIELDS},
            "changed_fields": changed_fields,
            "changed": True,
            "deleted": deleted,
            "added": True,
            "search_text": _build_search_text(current_values, deleted=deleted, added=True),
        }
        rows.append(row)

    if session_dir is not None:
        cache_path = _sheet_cache_path(session_dir)
        cache_path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    return rows


def filter_run_rows(rows: list[dict[str, Any]], query: str = "", changed_only: bool = False) -> list[dict[str, Any]]:
    """Фильтрует строки редактора по тексту и флагу changed_only."""
    q = (query or "").strip().lower()
    out: list[dict[str, Any]] = []
    for row in rows:
        if changed_only and not row.get("changed"):
            continue
        if q and q not in str(row.get("search_text", "")):
            continue
        out.append(row)
    return out


def row_snapshot_by_excel_row(run_xlsx: Path, excel_row: int) -> dict[str, Any] | None:
    """Возвращает исходное состояние одной строки РУН."""
    rows = read_run_rows(run_xlsx)
    for row in rows:
        if int(row.get("excel_row", -1) or -1) == int(excel_row):
            return row
    return None


def add_overlay_row(session_dir: Path, run_xlsx: Path, values: dict[str, Any] | None = None) -> tuple[dict[str, Any], str]:
    """Добавляет новую строку РУН в overlay и возвращает её row_id."""
    overlay = load_overlay(session_dir)
    added_rows = overlay.get(ADDED_ROWS_KEY, {})
    next_added_id = int(overlay.get(NEXT_ADDED_ID_KEY, 1) or 1)
    row_id = f"{ADDED_ROW_PREFIX}{next_added_id}"

    defaults = _default_new_row_values(run_xlsx)
    payload = {field: _normalize_cell_value(defaults.get(field, "")) for field in NEW_ROW_FIELDS}
    values = values if isinstance(values, dict) else {}
    for field in NEW_ROW_FIELDS:
        if field in values:
            payload[field] = _normalize_cell_value(values.get(field))

    clean_payload = {field: value for field, value in payload.items() if value != ""}
    added_rows[row_id] = clean_payload
    overlay[ADDED_ROWS_KEY] = added_rows
    overlay[NEXT_ADDED_ID_KEY] = next_added_id + 1
    save_overlay(session_dir, overlay)
    clear_overlay_artifacts(session_dir)
    return overlay, row_id


def update_overlay_row(session_dir: Path, run_xlsx: Path, row_id: Any, values: dict[str, Any]) -> dict[str, Any]:
    """Пересчитывает overlay для одной строки по полному набору введённых значений."""
    row_key = _normalize_row_id(row_id)
    overlay = load_overlay(session_dir)

    if _is_added_row_id(row_key):
        added_rows = overlay.get(ADDED_ROWS_KEY, {})
        previous_overlay = added_rows.get(row_key, {}) if isinstance(added_rows.get(row_key, {}), dict) else {}
        new_overlay: dict[str, str] = {}
        for field in NEW_ROW_FIELDS:
            new_value = _normalize_cell_value((values or {}).get(field, previous_overlay.get(field, "")))
            if new_value != "":
                new_overlay[field] = new_value
        if _overlay_row_deleted(previous_overlay):
            new_overlay[DELETED_FLAG] = "1"
        if new_overlay:
            added_rows[row_key] = new_overlay
        else:
            added_rows.pop(row_key, None)
        overlay[ADDED_ROWS_KEY] = added_rows
        save_overlay(session_dir, overlay)
        clear_overlay_artifacts(session_dir)
        return overlay

    snapshot = row_snapshot_by_excel_row(run_xlsx, int(row_key))
    if snapshot is None:
        raise ValueError(f"excel row {row_key} not found")

    original = snapshot.get("original", {}) if isinstance(snapshot, dict) else {}
    rows = overlay.get("rows", {})
    previous_overlay = rows.get(row_key, {}) if isinstance(rows.get(row_key, {}), dict) else {}

    new_overlay: dict[str, str] = {}
    for field in EDITABLE_FIELDS:
        original_value = _normalize_cell_value(original.get(field, ""))
        new_value = _normalize_cell_value((values or {}).get(field, original_value))
        if new_value != original_value:
            new_overlay[field] = new_value

    if _overlay_row_deleted(previous_overlay):
        new_overlay[DELETED_FLAG] = "1"

    if new_overlay:
        rows[row_key] = new_overlay
    else:
        rows.pop(row_key, None)
    overlay["rows"] = rows
    save_overlay(session_dir, overlay)
    clear_overlay_artifacts(session_dir)
    return overlay


def revert_overlay_row(session_dir: Path, row_id: Any) -> dict[str, Any]:
    """Удаляет все overlay-правки для одной строки."""
    row_key = _normalize_row_id(row_id)
    overlay = load_overlay(session_dir)
    if _is_added_row_id(row_key):
        added_rows = overlay.get(ADDED_ROWS_KEY, {})
        added_rows.pop(row_key, None)
        overlay[ADDED_ROWS_KEY] = added_rows
    else:
        rows = overlay.get("rows", {})
        rows.pop(row_key, None)
        overlay["rows"] = rows
    save_overlay(session_dir, overlay)
    clear_overlay_artifacts(session_dir)
    return overlay


def delete_overlay_row(session_dir: Path, row_id: Any) -> dict[str, Any]:
    """Помечает строку РУН как удалённую в overlay."""
    row_key = _normalize_row_id(row_id)
    overlay = load_overlay(session_dir)
    bucket_key = ADDED_ROWS_KEY if _is_added_row_id(row_key) else "rows"
    bucket = overlay.get(bucket_key, {})
    row_overlay = bucket.get(row_key, {}) if isinstance(bucket.get(row_key, {}), dict) else {}
    row_overlay[DELETED_FLAG] = "1"
    bucket[row_key] = row_overlay
    overlay[bucket_key] = bucket
    save_overlay(session_dir, overlay)
    clear_overlay_artifacts(session_dir)
    return overlay


def restore_overlay_row(session_dir: Path, row_id: Any) -> dict[str, Any]:
    """Снимает признак удаления со строки РУН, сохраняя остальные правки."""
    row_key = _normalize_row_id(row_id)
    overlay = load_overlay(session_dir)
    bucket_key = ADDED_ROWS_KEY if _is_added_row_id(row_key) else "rows"
    bucket = overlay.get(bucket_key, {})
    row_overlay = bucket.get(row_key, {}) if isinstance(bucket.get(row_key, {}), dict) else {}
    row_overlay.pop(DELETED_FLAG, None)
    if row_overlay:
        bucket[row_key] = row_overlay
    else:
        bucket.pop(row_key, None)
    overlay[bucket_key] = bucket
    save_overlay(session_dir, overlay)
    clear_overlay_artifacts(session_dir)
    return overlay


def revert_overlay_all(session_dir: Path) -> dict[str, Any]:
    """Удаляет все overlay-правки текущей сессии."""
    overlay = _empty_overlay()
    save_overlay(session_dir, overlay)
    clear_overlay_artifacts(session_dir)
    return overlay


def _excel_value_for_field(field: str, value: Any) -> Any:
    text = _normalize_cell_value(value)
    if field in NUMERIC_FIELDS:
        if text == "":
            return 0
        try:
            number = float(text.replace(",", "."))
            return int(number) if number.is_integer() else number
        except Exception:
            return text
    return text


def _freeze_formula_results_from_original(original_path: Path, workbook: Any) -> None:
    """Заменяет формулы на их текущие вычисленные значения только для внутреннего parser-friendly файла."""
    original_formula = load_workbook(original_path, data_only=False)
    original_values = load_workbook(original_path, data_only=True)
    try:
        for sheet_name in workbook.sheetnames:
            if sheet_name not in original_formula.sheetnames or sheet_name not in original_values.sheetnames:
                continue
            ws_formula = original_formula[sheet_name]
            ws_values = original_values[sheet_name]
            ws_target = workbook[sheet_name]
            for row in ws_formula.iter_rows():
                for cell in row:
                    if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                        continue
                    ws_target.cell(cell.row, cell.column).value = ws_values.cell(cell.row, cell.column).value
    finally:
        original_formula.close()
        original_values.close()


def _copy_row_format(ws: Any, src_row: int, dst_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = copy(src.number_format)
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.protection:
            dst.protection = copy(src.protection)

    if src_row in ws.row_dimensions:
        src_dim = ws.row_dimensions[src_row]
        dst_dim = ws.row_dimensions[dst_row]
        dst_dim.height = src_dim.height
        dst_dim.hidden = src_dim.hidden
        dst_dim.outlineLevel = src_dim.outlineLevel
        dst_dim.outline_level = src_dim.outline_level
        dst_dim.collapsed = src_dim.collapsed
        dst_dim.thickTop = src_dim.thickTop
        dst_dim.thickBot = src_dim.thickBot


def _copy_row_template(ws: Any, src_row: int, dst_row: int) -> None:
    """Копирует стиль строки и переносит формулы со сдвигом, не дублируя чужие значения."""
    _copy_row_format(ws, src_row, dst_row)
    for col in range(1, ws.max_column + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        value = src.value
        if isinstance(value, str) and value.startswith('='):
            try:
                dst.value = Translator(value, origin=src.coordinate).translate_formula(dst.coordinate)
            except Exception:
                dst.value = value
        else:
            dst.value = None


def _find_summary_row(ws: Any) -> int:
    for row_idx in range(ws.max_row, DATA_START_ROW - 1, -1):
        cell_value = ws.cell(row_idx, 39).value
        if isinstance(cell_value, str) and 'SUBTOTAL' in cell_value.upper():
            return row_idx
    return ws.max_row


def _set_recalc_flags(workbook: Any) -> None:
    try:
        workbook.calculation.calcMode = 'auto'
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
    except Exception:
        pass


def _rewrite_summary_row_formulas(ws: Any, summary_row: int) -> None:
    data_last_row = max(DATA_START_ROW, summary_row - 1)

    for col in range(39, 59):
        letter = get_column_letter(col)
        ws.cell(summary_row, col).value = f'=SUBTOTAL(9,{letter}{DATA_START_ROW}:{letter}{data_last_row})'

    ws.cell(summary_row, 59).value = f'=SUBTOTAL(9,BF{DATA_START_ROW}:BF{data_last_row})-BF{summary_row}'

    for col in range(60, 66):
        letter = get_column_letter(col)
        ws.cell(summary_row, col).value = f'=SUBTOTAL(9,{letter}{DATA_START_ROW}:{letter}{data_last_row})'

    ws.cell(summary_row, 66).value = f'=BG{summary_row}-BH{summary_row}-BJ{summary_row}-BK{summary_row}-BM{summary_row}'


def _rewrite_summary_refs_in_header(ws: Any, summary_row: int) -> None:
    ws['A1'] = '=BF1'
    for col in list(range(39, 59)) + list(range(60, 66)):
        letter = get_column_letter(col)
        ws.cell(1, col).value = f'={letter}{summary_row}'


def _apply_row_values(ws: Any, columns: dict[str, dict[str, Any]], excel_row: int, row_data: dict[str, Any], fields: list[str]) -> None:
    for field in fields:
        spec = columns.get(field)
        if spec is None:
            continue
        ws.cell(excel_row, spec['excel_col']).value = _excel_value_for_field(field, row_data.get(field, ''))


def _insert_added_row(ws: Any, columns: dict[str, dict[str, Any]], insert_row: int, row_data: dict[str, Any]) -> None:
    ws.insert_rows(insert_row, 1)
    template_row = max(DATA_START_ROW, insert_row - 1)
    _copy_row_template(ws, template_row, insert_row)
    _apply_row_values(ws, columns, insert_row, row_data, NEW_ROW_FIELDS)


def materialize_edited_run(run_xlsx: Path, session_dir: Path) -> Path:
    """Собирает run_edited.xlsx из исходного файла и overlay-правок с сохранением формул и оформления."""
    run_xlsx = Path(run_xlsx)
    out_path = _edited_run_path(session_dir)
    overlay = load_overlay(session_dir)
    rows = overlay.get("rows", {})
    added_rows = overlay.get(ADDED_ROWS_KEY, {})

    shutil.copyfile(run_xlsx, out_path)
    if not rows and not added_rows:
        return out_path

    sheet, columns, _ = _find_columns(run_xlsx)
    workbook = load_workbook(out_path)
    _set_recalc_flags(workbook)
    if sheet not in workbook.sheetnames:
        raise ValueError(f"sheet {sheet!r} not found in {run_xlsx}")
    ws = workbook[sheet]

    for row_key, changes in rows.items():
        try:
            excel_row = int(row_key)
        except Exception:
            continue
        if not isinstance(changes, dict):
            continue
        if excel_row < 1 or excel_row > ws.max_row:
            continue

        if _overlay_row_deleted(changes):
            for field in DELETE_CLEAR_FIELDS:
                spec = columns.get(field)
                if spec is None:
                    continue
                ws.cell(excel_row, spec['excel_col']).value = _cell_value_for_deleted(field)
            continue

        for field, value in changes.items():
            if field == DELETED_FLAG:
                continue
            spec = columns.get(field)
            if spec is None:
                continue
            ws.cell(excel_row, spec['excel_col']).value = _excel_value_for_field(field, value)

    summary_row = _find_summary_row(ws)
    live_added_rows = [row_data for _, row_data in sorted(added_rows.items(), key=lambda x: x[0]) if isinstance(row_data, dict) and not _overlay_row_deleted(row_data)]
    for row_data in live_added_rows:
        _insert_added_row(ws, columns, summary_row, row_data)
        summary_row += 1

    _rewrite_summary_row_formulas(ws, summary_row)
    _rewrite_summary_refs_in_header(ws, summary_row)

    workbook.save(out_path)
    workbook.close()
    return out_path


def materialize_effective_run(run_xlsx: Path, session_dir: Path) -> Path:
    """Собирает внутренний файл для rebuild: c замороженными значениями формул, удобный для pandas-парсера."""
    run_xlsx = Path(run_xlsx)
    out_path = _effective_materialized_run_path(session_dir)
    overlay = load_overlay(session_dir)
    rows = overlay.get("rows", {})
    added_rows = overlay.get(ADDED_ROWS_KEY, {})

    shutil.copyfile(run_xlsx, out_path)
    workbook = load_workbook(out_path)
    _freeze_formula_results_from_original(run_xlsx, workbook)

    if not rows and not added_rows:
        workbook.save(out_path)
        workbook.close()
        return out_path

    sheet, columns, _ = _find_columns(run_xlsx)
    if sheet not in workbook.sheetnames:
        raise ValueError(f"sheet {sheet!r} not found in {run_xlsx}")
    ws = workbook[sheet]

    for row_key, changes in rows.items():
        try:
            excel_row = int(row_key)
        except Exception:
            continue
        if not isinstance(changes, dict):
            continue
        if excel_row < 1 or excel_row > ws.max_row:
            continue

        if _overlay_row_deleted(changes):
            for field in DELETE_CLEAR_FIELDS:
                spec = columns.get(field)
                if spec is None:
                    continue
                ws.cell(excel_row, spec['excel_col']).value = _cell_value_for_deleted(field)
            continue

        for field, value in changes.items():
            if field == DELETED_FLAG:
                continue
            spec = columns.get(field)
            if spec is None:
                continue
            ws.cell(excel_row, spec['excel_col']).value = _excel_value_for_field(field, value)

    for _, row_data in sorted(added_rows.items(), key=lambda x: x[0]):
        if not isinstance(row_data, dict) or _overlay_row_deleted(row_data):
            continue
        dst_row = ws.max_row + 1
        template_row = max(DATA_START_ROW, dst_row - 1)
        _copy_row_format(ws, template_row, dst_row)
        for col in range(1, ws.max_column + 1):
            ws.cell(dst_row, col).value = None
        _apply_row_values(ws, columns, dst_row, row_data, NEW_ROW_FIELDS)

    workbook.save(out_path)
    workbook.close()
    return out_path


def effective_run_path(session_dir: Path) -> Path:
    """Возвращает путь к внутреннему parser-friendly файлу РУН для расчёта."""
    original = session_dir / "run.xlsx"
    if not original.exists():
        return original
    overlay_stats = get_overlay_stats(session_dir)
    if not overlay_stats["has_overlay"]:
        return original
    effective = _effective_materialized_run_path(session_dir)
    if effective.exists():
        return effective
    return materialize_effective_run(original, session_dir)


def ensure_run_edited_download(session_dir: Path) -> Path:
    """Готовит run_edited.xlsx для скачивания независимо от наличия правок."""
    original = session_dir / "run.xlsx"
    if not original.exists():
        return original
    return materialize_edited_run(original, session_dir)
