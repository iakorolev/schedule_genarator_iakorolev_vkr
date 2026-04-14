"""Сравнение автоматически построенного расписания с ручным эталоном."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

from .normalize import clean_text, normalize_day, normalize_disc, normalize_time, normalize_week_type, teacher_lastname, _txt

WEEK_MARK_RE = re.compile(r"^\[(Ч|З|В|Н)\]\s*", re.I)


def _normalize_teacher_header(value: Any) -> str:
    """Очищает имя преподавателя из заголовка ручного или автоматического файла."""
    s = clean_text(value)
    if not s:
        return ""
    s = s.split("\n")[0].strip()
    s = re.sub(r"\s+", " ", s)
    return s


def _extract_disc(text: str) -> str:
    """Извлекает и нормализует дисциплину из текста ячейки."""
    s = clean_text(text).strip("/;,- ")
    if not s:
        return ""
    s = WEEK_MARK_RE.sub("", s).strip()
    head = s.split(";", 1)[0].strip(" /-")
    return normalize_disc(head)


def _split_program_entries(cell_text: Any) -> list[dict[str, str]]:
    """Разделяет содержимое ячейки timetable_by_teachers на отдельные записи."""
    raw = _txt(cell_text)
    if not raw:
        return []
    entries: list[dict[str, str]] = []
    for line in str(raw).splitlines():
        line = clean_text(line)
        if not line:
            continue
        week = ""
        m = WEEK_MARK_RE.match(line)
        if m:
            week = m.group(1).upper()
            line = WEEK_MARK_RE.sub("", line).strip()
        entries.append({
            "raw_text": line,
            "week": week,
            "disc_key": _extract_disc(line),
        })
    if not entries:
        entries.append({"raw_text": clean_text(raw), "week": "", "disc_key": _extract_disc(raw)})
    return entries


def read_reference_teacher_timetable(reference_path: Path) -> pd.DataFrame:
    """Читает ручной эталонный файл преподавательского расписания."""
    wb = openpyxl.load_workbook(reference_path, data_only=True)
    rows: list[dict[str, Any]] = []
    for sheet_name in wb.sheetnames:
        if not str(sheet_name).strip().isdigit():
            continue
        ws = wb[sheet_name]
        if ws.max_row < 6 or ws.max_column < 5:
            continue
        teacher_cols: list[tuple[int, str, str]] = []
        for col in range(5, ws.max_column + 1):
            teacher = _normalize_teacher_header(ws.cell(3, col).value)
            if not teacher:
                continue
            teacher_cols.append((col, teacher, teacher_lastname(teacher)))
        if not teacher_cols:
            continue

        current_day = ""
        current_pair = ""
        current_time = ""
        for row in range(6, ws.max_row + 1):
            day_val = _txt(ws.cell(row, 1).value)
            pair_val = ws.cell(row, 2).value
            time_val = _txt(ws.cell(row, 3).value)
            week = normalize_week_type(ws.cell(row, 4).value)

            if day_val:
                current_day = normalize_day(day_val)
            if pair_val not in (None, ""):
                current_pair = str(int(pair_val)) if isinstance(pair_val, (int, float)) and float(pair_val).is_integer() else str(pair_val)
            if time_val:
                current_time = normalize_time(time_val)

            if not current_day or not current_pair:
                continue

            for col, teacher, tkey in teacher_cols:
                cell = _txt(ws.cell(row, col).value)
                if not cell:
                    continue
                rows.append({
                    "sheet": str(sheet_name),
                    "teacher": teacher,
                    "teacher_key": tkey,
                    "day": current_day,
                    "pair": current_pair,
                    "time": current_time,
                    "week": week,
                    "raw_text": cell,
                    "disc_key": _extract_disc(cell),
                })
    return pd.DataFrame(rows)


def read_generated_teacher_timetable(generated_path: Path) -> pd.DataFrame:
    """Читает автоматически построенную таблицу timetable_by_teachers.xlsx."""
    df = pd.read_excel(generated_path).fillna("")
    if len(df) == 0:
        return pd.DataFrame(columns=["teacher", "teacher_key", "day", "pair", "time", "week", "raw_text", "disc_key"])
    week_col = None
    teacher_start_idx = 3
    if len(df.columns) >= 4 and str(df.columns[3]).strip().lower() in {"week_type", "week", "ч-з", "ч/з"}:
        week_col = df.columns[3]
        teacher_start_idx = 4
    teacher_cols = [c for c in df.columns[teacher_start_idx:] if _txt(c)]
    rows: list[dict[str, Any]] = []
    for _, r in df.iterrows():
        day = normalize_day(r.get(df.columns[0], ""))
        pair_raw = r.get(df.columns[1], "")
        pair = str(int(pair_raw)) if isinstance(pair_raw, (int, float)) and str(pair_raw) != "" and float(pair_raw).is_integer() else str(_txt(pair_raw))
        time = normalize_time(r.get(df.columns[2], ""))
        row_week = normalize_week_type(r.get(week_col, "")) if week_col else ""
        if not day or not pair:
            continue
        for teacher in teacher_cols:
            teacher_name = _normalize_teacher_header(teacher)
            if not teacher_name:
                continue
            cell = r.get(teacher, "")
            if _txt(cell) == "":
                continue
            for entry in _split_program_entries(cell):
                rows.append({
                    "teacher": teacher_name,
                    "teacher_key": teacher_lastname(teacher_name),
                    "day": day,
                    "pair": pair,
                    "time": time,
                    "week": normalize_week_type(entry.get("week", "")) or row_week,
                    "raw_text": entry.get("raw_text", ""),
                    "disc_key": entry.get("disc_key", ""),
                })
    return pd.DataFrame(rows)


def _aggregate_records(df: pd.DataFrame, source: str) -> pd.DataFrame:
    """Агрегирует записи по преподавателю, слоту и типу недели."""
    if df is None or len(df) == 0:
        return pd.DataFrame(columns=["teacher_key", "teacher", "day", "pair", "time", "week", f"{source}_count", f"{source}_texts", f"{source}_disc_keys"])
    grouped = (
        df.groupby(["teacher_key", "day", "pair", "time", "week"], as_index=False)
        .agg(
            teacher=("teacher", "first"),
            record_count=("raw_text", "count"),
            raw_texts=("raw_text", lambda s: "\n".join(dict.fromkeys([_txt(x) for x in s if _txt(x)]))),
            disc_keys=("disc_key", lambda s: ", ".join(sorted(dict.fromkeys([_txt(x) for x in s if _txt(x)])))),
        )
    )
    grouped = grouped.rename(columns={
        "teacher": f"{source}_teacher",
        "record_count": f"{source}_count",
        "raw_texts": f"{source}_texts",
        "disc_keys": f"{source}_disc_keys",
        "week": f"{source}_week",
    })
    return grouped


def _weeks_compatible(a: Any, b: Any) -> bool:
    aw = normalize_week_type(a)
    bw = normalize_week_type(b)
    if not aw or not bw:
        return True
    return aw == bw


def _build_slot_compare(ref_common: pd.DataFrame, gen_common: pd.DataFrame) -> pd.DataFrame:
    """Строит сравнение слотов с учётом типа недели и wildcard-логики для пустой недели."""
    rows: list[dict[str, Any]] = []
    keys = set()
    if ref_common is not None and len(ref_common) > 0:
        keys.update(tuple(x) for x in ref_common[["teacher_key", "day", "pair", "time"]].drop_duplicates().itertuples(index=False, name=None))
    if gen_common is not None and len(gen_common) > 0:
        keys.update(tuple(x) for x in gen_common[["teacher_key", "day", "pair", "time"]].drop_duplicates().itertuples(index=False, name=None))

    for teacher_key, day, pair, time in sorted(keys):
        ref_slot = ref_common[(ref_common["teacher_key"] == teacher_key) & (ref_common["day"] == day) & (ref_common["pair"] == pair) & (ref_common["time"] == time)].copy()
        gen_slot = gen_common[(gen_common["teacher_key"] == teacher_key) & (gen_common["day"] == day) & (gen_common["pair"] == pair) & (gen_common["time"] == time)].copy()

        ref_records = ref_slot.to_dict(orient="records")
        gen_records = gen_slot.to_dict(orient="records")
        used_gen: set[int] = set()

        for ref_rec in ref_records:
            match_idx = None
            for gi, gen_rec in enumerate(gen_records):
                if gi in used_gen:
                    continue
                if _weeks_compatible(ref_rec.get("manual_week"), gen_rec.get("program_week")):
                    match_idx = gi
                    break
            if match_idx is not None:
                gen_rec = gen_records[match_idx]
                used_gen.add(match_idx)
                rows.append({
                    "teacher_key": teacher_key,
                    "day": day,
                    "pair": pair,
                    "time": time,
                    "manual_teacher": ref_rec.get("manual_teacher", ""),
                    "program_teacher": gen_rec.get("program_teacher", ""),
                    "manual_count": ref_rec.get("manual_count"),
                    "program_count": gen_rec.get("program_count"),
                    "manual_texts": ref_rec.get("manual_texts", ""),
                    "program_texts": gen_rec.get("program_texts", ""),
                    "manual_disc_keys": ref_rec.get("manual_disc_keys", ""),
                    "program_disc_keys": gen_rec.get("program_disc_keys", ""),
                    "manual_weeks": ref_rec.get("manual_week", ""),
                    "program_weeks": gen_rec.get("program_week", ""),
                    "week_overlap": True,
                    "presence_status": "both",
                })
            else:
                rows.append({
                    "teacher_key": teacher_key,
                    "day": day,
                    "pair": pair,
                    "time": time,
                    "manual_teacher": ref_rec.get("manual_teacher", ""),
                    "program_teacher": "",
                    "manual_count": ref_rec.get("manual_count"),
                    "program_count": None,
                    "manual_texts": ref_rec.get("manual_texts", ""),
                    "program_texts": "",
                    "manual_disc_keys": ref_rec.get("manual_disc_keys", ""),
                    "program_disc_keys": "",
                    "manual_weeks": ref_rec.get("manual_week", ""),
                    "program_weeks": "",
                    "week_overlap": False,
                    "presence_status": "manual_only",
                })

        for gi, gen_rec in enumerate(gen_records):
            if gi in used_gen:
                continue
            rows.append({
                "teacher_key": teacher_key,
                "day": day,
                "pair": pair,
                "time": time,
                "manual_teacher": "",
                "program_teacher": gen_rec.get("program_teacher", ""),
                "manual_count": None,
                "program_count": gen_rec.get("program_count"),
                "manual_texts": "",
                "program_texts": gen_rec.get("program_texts", ""),
                "manual_disc_keys": "",
                "program_disc_keys": gen_rec.get("program_disc_keys", ""),
                "manual_weeks": "",
                "program_weeks": gen_rec.get("program_week", ""),
                "week_overlap": False,
                "presence_status": "program_only",
            })
    return pd.DataFrame(rows)

def _disc_set(value: Any) -> set[str]:
    return {x.strip() for x in str(_txt(value)).split(",") if x and str(x).strip()}


def compare_teacher_timetables(reference_path: Path, generated_path: Path, out_dir: Path) -> dict[str, Any]:
    """Сравнивает ручной эталон и автоматически построенную преподавательскую сетку."""
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    ref_df = read_reference_teacher_timetable(Path(reference_path))
    gen_df = read_generated_teacher_timetable(Path(generated_path))

    ref_aggr = _aggregate_records(ref_df, "manual")
    gen_aggr = _aggregate_records(gen_df, "program")

    teacher_ref = (
        ref_aggr.groupby("teacher_key", as_index=False)
        .agg(manual_teacher=("manual_teacher", "first"), manual_slots=("pair", "count"))
        if len(ref_aggr) > 0 else pd.DataFrame(columns=["teacher_key", "manual_teacher", "manual_slots"])
    )
    teacher_gen = (
        gen_aggr.groupby("teacher_key", as_index=False)
        .agg(program_teacher=("program_teacher", "first"), program_slots=("pair", "count"))
        if len(gen_aggr) > 0 else pd.DataFrame(columns=["teacher_key", "program_teacher", "program_slots"])
    )
    teacher_cmp = teacher_ref.merge(teacher_gen, on="teacher_key", how="outer")
    teacher_cmp["manual_slots"] = teacher_cmp.get("manual_slots", 0).fillna(0).astype(int)
    teacher_cmp["program_slots"] = teacher_cmp.get("program_slots", 0).fillna(0).astype(int)
    teacher_cmp["slot_delta"] = teacher_cmp["program_slots"] - teacher_cmp["manual_slots"]
    teacher_cmp["teacher_status"] = teacher_cmp.apply(
        lambda r: "common" if _txt(r.get("manual_teacher")) and _txt(r.get("program_teacher")) else ("manual_only" if _txt(r.get("manual_teacher")) else "program_only"),
        axis=1,
    )
    teacher_cmp = teacher_cmp.sort_values(["teacher_status", "teacher_key", "manual_teacher", "program_teacher"]).reset_index(drop=True)

    common_keys = set(teacher_cmp.loc[teacher_cmp["teacher_status"] == "common", "teacher_key"].tolist())
    ref_common = ref_aggr[ref_aggr["teacher_key"].isin(common_keys)].copy() if len(ref_aggr) > 0 else ref_aggr.copy()
    gen_common = gen_aggr[gen_aggr["teacher_key"].isin(common_keys)].copy() if len(gen_aggr) > 0 else gen_aggr.copy()

    slot_cmp = _build_slot_compare(ref_common, gen_common)
    if len(slot_cmp) == 0:
        slot_cmp = pd.DataFrame(columns=[
            "teacher_key", "day", "pair", "time", "manual_teacher", "program_teacher", "manual_count", "program_count",
            "manual_texts", "program_texts", "manual_disc_keys", "program_disc_keys", "manual_weeks", "program_weeks",
            "week_overlap", "presence_status", "discipline_overlap",
        ])
    else:
        slot_cmp["manual_disc_set"] = slot_cmp["manual_disc_keys"].apply(_disc_set)
        slot_cmp["program_disc_set"] = slot_cmp["program_disc_keys"].apply(_disc_set)
        slot_cmp["discipline_overlap"] = slot_cmp.apply(
            lambda r: bool(r["manual_disc_set"] & r["program_disc_set"]) if r["presence_status"] == "both" else False,
            axis=1,
        )
        slot_cmp["manual_teacher"] = slot_cmp.get("manual_teacher", "").fillna("")
        slot_cmp["program_teacher"] = slot_cmp.get("program_teacher", "").fillna("")
        slot_cmp["teacher_display"] = slot_cmp.apply(lambda r: _txt(r.get("manual_teacher")) or _txt(r.get("program_teacher")), axis=1)
        slot_cmp = slot_cmp.drop(columns=[c for c in ["manual_disc_set", "program_disc_set"] if c in slot_cmp.columns])
        slot_cmp = slot_cmp.sort_values(["teacher_display", "day", "pair", "time", "manual_weeks", "program_weeks"]).reset_index(drop=True)

    mismatch_presence = slot_cmp[slot_cmp["presence_status"] != "both"].copy() if len(slot_cmp) > 0 else slot_cmp.copy()
    mismatch_disc = slot_cmp[(slot_cmp["presence_status"] == "both") & (~slot_cmp["discipline_overlap"])].copy() if len(slot_cmp) > 0 else slot_cmp.copy()

    teacher_quality = None
    if len(slot_cmp) > 0:
        teacher_quality = (
            slot_cmp.groupby(["teacher_key", "teacher_display"], as_index=False)
            .agg(
                compared_slots=("pair", "count"),
                both_present=("presence_status", lambda s: int(sum(x == "both" for x in s))),
                discipline_overlap_slots=("discipline_overlap", lambda s: int(sum(bool(x) for x in s))),
                manual_only_slots=("presence_status", lambda s: int(sum(x == "manual_only" for x in s))),
                program_only_slots=("presence_status", lambda s: int(sum(x == "program_only" for x in s))),
            )
        )
        teacher_quality["presence_match_rate"] = teacher_quality.apply(
            lambda r: round((r["both_present"] / r["compared_slots"]) if r["compared_slots"] else 0.0, 4), axis=1
        )
        teacher_quality["discipline_overlap_rate"] = teacher_quality.apply(
            lambda r: round((r["discipline_overlap_slots"] / r["both_present"]) if r["both_present"] else 0.0, 4), axis=1
        )
        teacher_quality = teacher_quality.sort_values(["discipline_overlap_rate", "presence_match_rate", "teacher_display"]).reset_index(drop=True)
    else:
        teacher_quality = pd.DataFrame(columns=["teacher_key", "teacher_display", "compared_slots", "both_present", "discipline_overlap_slots", "manual_only_slots", "program_only_slots", "presence_match_rate", "discipline_overlap_rate"])

    manual_teacher_count = int((teacher_cmp["teacher_status"] != "program_only").sum()) if len(teacher_cmp) > 0 else 0
    program_teacher_count = int((teacher_cmp["teacher_status"] != "manual_only").sum()) if len(teacher_cmp) > 0 else 0
    common_teacher_count = int((teacher_cmp["teacher_status"] == "common").sum()) if len(teacher_cmp) > 0 else 0
    manual_slots_common = int(ref_common.shape[0]) if ref_common is not None else 0
    program_slots_common = int(gen_common.shape[0]) if gen_common is not None else 0
    slot_both = int((slot_cmp["presence_status"] == "both").sum()) if len(slot_cmp) > 0 else 0
    slot_disc_overlap = int(((slot_cmp["presence_status"] == "both") & (slot_cmp["discipline_overlap"])).sum()) if len(slot_cmp) > 0 else 0

    summary = pd.DataFrame([
        {"metric": "manual_teacher_count", "value": manual_teacher_count},
        {"metric": "program_teacher_count", "value": program_teacher_count},
        {"metric": "common_teacher_count", "value": common_teacher_count},
        {"metric": "manual_only_teacher_count", "value": int((teacher_cmp["teacher_status"] == "manual_only").sum()) if len(teacher_cmp) > 0 else 0},
        {"metric": "program_only_teacher_count", "value": int((teacher_cmp["teacher_status"] == "program_only").sum()) if len(teacher_cmp) > 0 else 0},
        {"metric": "manual_slots_common_teachers", "value": manual_slots_common},
        {"metric": "program_slots_common_teachers", "value": program_slots_common},
        {"metric": "slot_presence_matches", "value": slot_both},
        {"metric": "slot_discipline_overlap", "value": slot_disc_overlap},
        {"metric": "presence_recall_vs_manual", "value": round(slot_both / manual_slots_common, 4) if manual_slots_common else 0.0},
        {"metric": "presence_precision_vs_program", "value": round(slot_both / program_slots_common, 4) if program_slots_common else 0.0},
        {"metric": "discipline_overlap_rate_on_both", "value": round(slot_disc_overlap / slot_both, 4) if slot_both else 0.0},
    ])

    summary_path = out_dir / "compare_summary.xlsx"
    teachers_path = out_dir / "compare_teachers.xlsx"
    slots_path = out_dir / "compare_slots.xlsx"
    mism_presence_path = out_dir / "compare_presence_mismatches.xlsx"
    mism_disc_path = out_dir / "compare_discipline_mismatches.xlsx"
    quality_path = out_dir / "compare_teacher_quality.xlsx"

    summary.to_excel(summary_path, index=False)
    teacher_cmp.to_excel(teachers_path, index=False)
    slot_cmp.to_excel(slots_path, index=False)
    mismatch_presence.to_excel(mism_presence_path, index=False)
    mismatch_disc.to_excel(mism_disc_path, index=False)
    teacher_quality.to_excel(quality_path, index=False)

    return {
        "summary": {
            row["metric"]: row["value"] for row in summary.to_dict(orient="records")
        },
        "files": {
            "compare_summary": str(summary_path),
            "compare_teachers": str(teachers_path),
            "compare_slots": str(slots_path),
            "compare_presence_mismatches": str(mism_presence_path),
            "compare_discipline_mismatches": str(mism_disc_path),
            "compare_teacher_quality": str(quality_path),
        },
    }
